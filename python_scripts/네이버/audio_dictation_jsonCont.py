from crowdworks.commonUtil import *

import os
import shutil
import json
import uuid

from openpyxl import Workbook

ROOT_PATH = "/Users/myeonghyeonjeong/Documents/2019/09/네이버_음성챌린지/소스/"

COPY_PATH_16k = "/Users/myeonghyeonjeong/Documents/2019/08/네이버_음성받아쓰기_스크립트/소스/16k/"
COPY_PATH_script = "/Users/myeonghyeonjeong/Documents/2019/09/네이버_음성챌린지/txt/"
COPY_PATH_wav = "/Users/myeonghyeonjeong/Documents/2019/09/네이버_음성챌린지/wav/"


# 1. 파일 분류 16k.wav 만
def categorized_16k():
    for f in os.listdir(ROOT_PATH):
        if f.find('.16k.') != -1:
            shutil.copy2(ROOT_PATH + f, COPY_PATH_16k + f)


# 함수 실행
# categorized_16k()

# # 2. 파일 분류 .script 만
def categorized_script():
    for f in os.listdir(ROOT_PATH):
        if f.find('.txt') != -1:
            shutil.copy2(ROOT_PATH + f, COPY_PATH_script + f)


# # 함수 실행
# categorized_script()

# 3. 파일 분류 .wav 만
def categorized_wav():
    for f in os.listdir(ROOT_PATH):
        if f.find('.wav') != -1 and f.find('.16k.') == -1:
            shutil.copy2(ROOT_PATH + f, COPY_PATH_wav + f)


# 함수 실행
# categorized_wav()

# 4. jsonCont 에 사용할 엑셀 만들기
def jsonContGenerator():
    write_wb = Workbook()

    write_ws = write_wb.active

    write_ws['A1'] = '0'
    write_ws['B1'] = '0'
    write_ws['C1'] = '0'
    write_ws['D1'] = '0'

    row = 1
    idx = 0

    for f in os.listdir(ROOT_PATH):
        if f.find('.wav') != -1 and f.find('.16k.') == -1:
            row += 1
            idx += 1
            # print(f.split('.wav')[0])
            write_ws.cell(row, 1, f.split('.wav')[0]+'.wav')
            write_ws.cell(row, 2, f.split('.wav')[0]+'.wav')

            file = open("/Users/myeonghyeonjeong/Documents/2019/09/네이버_음성챌린지/txt/"+f.split('.wav')[0]+".txt",'r',encoding='utf-8')
            write_ws.cell(row, 3, file.read().replace('\n', ''))
            write_ws.cell(row, 4, idx)
            file.close()

    write_wb.save('/Users/myeonghyeonjeong/Documents/2019/09/네이버_음성챌린지/소스/jsonCont.xlsx')


jsonContGenerator()
