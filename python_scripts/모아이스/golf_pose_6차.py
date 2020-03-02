"""
[모아이스] 골프자세추출

# 6차 프로젝트아이디: 1872,1873,1874,1875,1876,1877,1878,1879,1880,1881
# 6차 태스크아이디 (139개): 3263,3264,3265,3266,3267,3268,3269,3270,3271,3272,3273,3274,3275,3276,3277,3278,3279,3280,3281,3282,3283,3284,3285,3286,3287,3288,3289,3290,3291,3292,3293,3294,3295,3296,3297,3298,3299,3300,3301,3302,3303,3304,3305,3306,3307,3308,3309,3310,3311,3312,3313,3314,3315,3316,3317,3318,3319,3320,3321,3322,3323,3324,3325,3326,3327,3328,3329,3330,3331,3332,3333,3334,3335,3336,3337,3338,3339,3340,3341,3342,3343,3344,3345,3346,3347,3348,3349,3350,3351,3352,3353,3354,3355,3356,3357,3358,3359,3360,3361,3362,3363,3364,3365,3366,3367,3368,3369,3370,3371,3372,3373,3374,3375,3376,3377,3378,3379,3380,3381,3382,3383,3384,3385,3386,3387,3388,3389,3390,3391,3392,3393,3394,3395,3396,3397,3398,3399,3400
19.09.18 - 1개 잘못 추가함. 138개가 맞음.
"""

from crowdworks.commonUtil import *

import os
import json
import shutil

from openpyxl import Workbook

from PIL import Image  # 이미지 사이즈 가져오기용 라이브러리
import pandas as pd  # csv 변환용 라이브러리

# 5차 프로젝트아이디: 1872,1873,1874,1875,1876,1877,1878,1879,1880,1881
TASK_IDS = [3263, 3264, 3265, 3266, 3267, 3268, 3269, 3270, 3271, 3272, 3273, 3274, 3275, 3276, 3277, 3278, 3279, 3280,
            3281, 3282, 3283, 3284, 3285, 3286, 3287, 3288, 3289, 3290, 3291, 3292, 3293, 3294, 3295, 3296, 3297, 3298,
            3299, 3300, 3301, 3302, 3303, 3304, 3305, 3306, 3307, 3308, 3309, 3310, 3311, 3312, 3313, 3314, 3315, 3316,
            3317, 3318, 3319, 3320, 3321, 3322, 3323, 3324, 3325, 3326, 3327, 3328, 3329, 3330, 3331, 3332, 3333, 3334,
            3335, 3336, 3337, 3338, 3339, 3340, 3341, 3342, 3343, 3344, 3345, 3346, 3347, 3348, 3349, 3350, 3351, 3352,
            3353, 3354, 3355, 3356, 3357, 3358, 3359, 3360, 3361, 3362, 3363, 3364, 3365, 3366, 3367, 3368, 3369, 3370,
            3371, 3372, 3373, 3374, 3375, 3376, 3377, 3378, 3379, 3380, 3381, 3382, 3383, 3384, 3385, 3386, 3387, 3388,
            3389, 3390, 3391, 3392, 3393, 3394, 3395, 3396, 3397, 3398, 3399, 3400]
FOLDER_CATEGORY = ["Front/", "Side/"]  # Front/ - 58 , Side/ - 80, Total - 138
FOLDER_FRONT = ['1806547534941746630', '1803674812494434928', '1625668272509984270', '1621687735214707044',
                '1587693811684770781', '1805804418181773826', '1802181603628958816', '1628037476160317852',
                '1798371761964880275', '1806616842266064221', '1583173035635986706', '1573138403297237227',
                '1621686899088667365', '1632398335988745826', '1804269683131473293', '1792748944846463946',
                '1621719645764443379', '1804089204293928604', '1626828654368482286', '1799366974481779206',
                '1617188553002068891', '1610999625547562363', '1793359937390197641', '1620344367615519386',
                '1589092637838444181', '1597701989256380568', '1795436944835395613', '1792593671293279843',
                '1593328689089009845', '1602213878404753300', '1794896568308457619', '1577973772521167511',
                '1626184595005274885', '1604453230363562627', '1633329445497406278', '1592780683163419276',
                '1578488245146445142', '1604248172208532628', '1622982617840528292', '1612375391325733685',
                '1610090059075076461', '1799977754716960573', '1625441849577503386', '1577355209880676355',
                '1633204786143498934', '1622356002153460847', '1794121962356950513', '1596334406178578336',
                '1794080785725974125', '1594196635406953814', '1797160180908928669', '1627530442393811534',
                '1604180669809206512', '1621687320658137097', '1801518738164582926', '1575391140520920918',
                '1795587053640908383', '1801378079068650494']
FOLDER_SIDE = ['1986160843138884566', '1979133999785130562', '1981056242957943943', '1979021913654495230',
               '1975346718796224557', '1980088262480673753', '1986426353302763799', '1978416292588428841',
               '1986942011236995295', '1980952515269902005', '1979894811541846088', '1983906806695900099',
               '1984922993322668780', '1980925512793721123', '1983107635428531003', '1986090910040330471',
               '1977634093991908149', '1982600157058064971', '1975359480041575285', '1986928288019633180',
               '1986280096385917008', '1984113506013738516', '1975396573895738319', '1982775749958388336',
               '1976602967311580164', '1987199477002771820', '1986553425085884491', '1984070264754769207',
               '1985957453779918733', '1987100080436856444', '1981359453891888059', '1976738684394203165',
               '1982395093240751551', '1981190465895356955', '1986294989119636962', '1984081627668930971',
               '1986599107230616293', '1985992870012993016', '1986204789177303120', '1982588816918325993',
               '1984679432154358776', '1977417702466556859', '1983125558662755880', '1982645226405144192',
               '1986559745507811062', '1985937857865623857', '1986870992400978191', '1985534975957918550',
               '1975572454517954619', '1983160056897041943', '1986310311910026351', '1977263225537260288',
               '1982018742444878661', '1984916445854799952', '1986878108532696064', '1986953879297120284',
               '1985587047655001063', '1986503010759955457', '1978073912584914528', '1987083665147807727',
               '1978897755310342559', '1978945098205873515', '1984129220604233046', '1985344336015358321',
               '1978912520862684514', '1984305323768267562', '1984184975319914982', '1981076025056107046',
               '1979667325980655599', '1986736942042542535', '1982363874247319563', '1982686685659224846',
               '1982394950784882923', '1975899343002294029', '1983857042855838172', '1986501567525147312',
               '1982640734389489377', '1983449311928950384', '1986973323293544018', '1980677574640107333']

EXT_PATH_ORIGIN = "/Users/myeonghyeonjeong/Desktop/DATA_EXT/모아이스/골프자세추출/"

import datetime

now = datetime.datetime.now()
year = str(now.year)
month = str(now.month).zfill(2)
day = str(now.day).zfill(2)

##### 6차 시작 #####
try:
    N_CHA = "labeled_data_1872-1881_" + year + month + day + "/"
    FOLDER_CATEGORY = ""
    PATH = EXT_PATH_ORIGIN + N_CHA
    os.mkdir(PATH)
    for category in FOLDER_CATEGORY:
        FOLDER_CATEGORY = category
        os.mkdir(PATH + FOLDER_CATEGORY)
except FileExistsError as e:
    print("##### 기존에 추출한 " + N_CHA + " 데이터를 모두 삭제하고 폴더 새로 만듦 #####")
    shutil.rmtree(PATH)
    os.mkdir(PATH)

    for category in FOLDER_CATEGORY:
        FOLDER_CATEGORY = category
        os.mkdir(PATH + FOLDER_CATEGORY)

for task_id in TASK_IDS:
    QUERY = "SELECT TPD.result_json, CFU.store_file_name, CFU.store_file_path FROM  TB_PRJ_DATA AS TPD INNER JOIN TB_MEMBER AS TM ON TM.member_id = TPD.work_user INNER JOIN TB_TMPL_IMG_LABEL_DATA AS LD ON TPD.data_idx  = LD.data_idx INNER JOIN CW_SOURCE AS CS ON CS.source_id = LD.src_idx INNER JOIN CW_FILE_UNPACKED AS CFU ON CFU.file_id = CS.file_id WHERE TPD.prj_idx IN (" + str(
        task_id) + ") AND (TPD.prog_state_cd='ALL_FINISHED' OR TPD.prog_state_cd='CHECK_END') ORDER BY store_file_name"
    RESULT = getDatabaseData(QUERY, "prd", "mhjeong", "cworks@34")

    try:
        FOLDER_NAME = RESULT[0][2].split('/')[6]

        if FOLDER_NAME in FOLDER_FRONT:
            EXT_PATH = EXT_PATH_ORIGIN + N_CHA + FOLDER_CATEGORY[0] + FOLDER_NAME + "/"
        if FOLDER_NAME in FOLDER_SIDE:
            EXT_PATH = EXT_PATH_ORIGIN + N_CHA + FOLDER_CATEGORY[1] + FOLDER_NAME + "/"

        try:
            os.mkdir(EXT_PATH)
        except FileExistsError:
            print("나면 안되는 오류인데 났음")
            break

        first_file = os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[0]

        if os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[0] == ".DS_Store":
            first_file = os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[1]

        # PIL 라이브러리에 Image 사용해 사이즈 가져오기
        image_path = "/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME + "/" + first_file
        image = Image.open(image_path)
        width, height = image.size

        # 2. 추출한 값들로 엑셀데이터 만들기
        row_idx = 1
        write_wb = Workbook()
        sheet1 = write_wb.active
        xlsx_file_name = EXT_PATH + FOLDER_NAME + ".xlsx"

        sheet1.cell(row=row_idx, column=1).value = "filename"
        sheet1.cell(row=row_idx, column=2).value = "width"
        sheet1.cell(row=row_idx, column=3).value = "height"
        sheet1.cell(row=row_idx, column=4).value = "class"
        sheet1.cell(row=row_idx, column=5).value = "x"
        sheet1.cell(row=row_idx, column=6).value = "y"

        for data in RESULT:
            result_json = json.loads(data[0])
            file_name = data[1]

            for category in result_json:
                for obj in result_json[category]['data']:
                    row_idx = row_idx + 1
                    sheet1.cell(row=row_idx, column=1).value = file_name
                    sheet1.cell(row=row_idx, column=2).value = width
                    sheet1.cell(row=row_idx, column=3).value = height
                    sheet1.cell(row=row_idx, column=4).value = obj['class']
                    sheet1.cell(row=row_idx, column=5).value = obj['x']
                    sheet1.cell(row=row_idx, column=6).value = obj['y']

        print(xlsx_file_name)
        write_wb.save(filename=xlsx_file_name)
        # / 2. 추출한 값들로 엑셀데이터 만들기

        # 3. csv 변환 후 기존 엑셀데이터 삭제
        data_xls = pd.read_excel(xlsx_file_name, 'Sheet', index_col=None)
        data_xls.to_csv(EXT_PATH + FOLDER_NAME + '.csv', encoding='utf-8')
        os.remove(xlsx_file_name)
        # / 3. csv 변환 후 기존 엑셀데이터 삭제
    except IndexError:
        print("IndexError")
        pass
##### 6차 끝 #####

# select U.store_file_path
#
# from  TB_PRJ_DATA     D left join TB_MEMBER M on M.member_id = D.work_user
#
# left join TB_TMPL_IMG_LABEL_DATA L on D.data_idx  = L.data_idx
# left join CW_SOURCE S on S.source_id = L.src_idx
# left join CW_FILE_UNPACKED U  on U.file_id = S.file_id
#
#
# where D.prj_idx in () and D.prog_state_cd="ALL_FINISHED";
