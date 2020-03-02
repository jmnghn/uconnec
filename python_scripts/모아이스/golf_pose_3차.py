"""
[모아이스] 골프자세추출

# 3차 프로젝트아이디: 1793, 1794, 1795, 1796, 1797, 1798, 1799, 1800, 1801, 1802
# 3차 태스크아이디 (100개): 2808,2809,2810,2811,2812,2813,2814,2815,2816,2817,2818,2819,2820,2821,2822,2823,2824,2825,2826,2827,2828,2829,2830,2831,2832,2833,2834,2835,2836,2837,2838,2839,2840,2841,2842,2843,2844,2845,2846,2847,2848,2849,2850,2851,2852,2853,2854,2855,2856,2857,2858,2859,2860,2861,2862,2863,2864,2865,2866,2867,2868,2869,2870,2871,2872,2873,2874,2875,2876,2877,2878,2879,2880,2881,2882,2883,2884,2885,2886,2887,2888,2889,2890,2891,2892,2893,2894,2895,2896,2897,2898,2899,2900,2901,2902,2903,2904,2905,2906,2907
"""

from crowdworks.commonUtil import *

import os
import json
import shutil

from openpyxl import Workbook

from PIL import Image  # 이미지 사이즈 가져오기용 라이브러리
import pandas as pd  # csv 변환용 라이브러리

# 3차 프로젝트아이디: 1793, 1794, 1795, 1796, 1797, 1798, 1799, 1800, 1801, 1802
TASK_IDS = [2808, 2809, 2810, 2811, 2812, 2813, 2814, 2815, 2816, 2817, 2818, 2819, 2820, 2821, 2822, 2823, 2824,
                  2825, 2826, 2827, 2828, 2829, 2830, 2831, 2832, 2833, 2834, 2835, 2836, 2837, 2838, 2839, 2840, 2841,
                  2842, 2843, 2844, 2845, 2846, 2847, 2848, 2849, 2850, 2851, 2852, 2853, 2854, 2855, 2856, 2857, 2858,
                  2859, 2860, 2861, 2862, 2863, 2864, 2865, 2866, 2867, 2868, 2869, 2870, 2871, 2872, 2873, 2874, 2875,
                  2876, 2877, 2878, 2879, 2880, 2881, 2882, 2883, 2884, 2885, 2886, 2887, 2888, 2889, 2890, 2891, 2892,
                  2893, 2894, 2895, 2896, 2897, 2898, 2899, 2900, 2901, 2902, 2903, 2904, 2905, 2906, 2907]
FOLDER_CATEGORY = ["Front/", "Side/", "Etc/"]  # Front/ - 20 , Side/ - 25 , Youtube/ - 55, Total - 100
FOLDER_FRONT = ['1987500002380904738', '1987473586462904627', '1987681557225377766', '1987450993985649770',
                      '1987635694768969008', '1987602139733007011', '1987390540836227377', '1987735919809124276',
                      '1987519781931533382', '1987389110361633843', '1987727376524972343', '1987590144877735928',
                      '1987318545591149235', '1987473494137854501', '1987284602944175615', '1987500481690501413',
                      '1987437937227572542', '1987636220600370585', '1987484069461524433', '1987701346632376131']
FOLDER_SIDE = ['1955235776821528499', '1957187763079287405', '1955903970653750525', '1955119143998421367',
                     '1957235036123445833', '1955616385414848498', '1955697345112091330', '1957001662405322656',
                     '1955834428128627187', '1954209140311366747', '1956618855830132127', '1955216614850101627',
                     '1956645163930955695', '1953747331938447820', '1955632101555846948', '1957168722926695764',
                     '1955529726772707935', '1955780025967223471', '1956667154750679677', '1954767580993921547',
                     '1955738567537165395', '1954211997437735402', '1956370975567553439', '1955106493532793215',
                     '1957019963253592683']
FOLDER_ETC = ['1921604959809087103', '1917885764546731179', '1934002190825894546', '1921598966736361690',
                    '1905619092851117689', '1321712923822849561', '1164985148276701703', '1927821122222778345',
                    '1936072649386419510', '1927849278459777123', '1935562722893407812', '1927830313771491901',
                    '1927839713097858218', '1920788912425301065', '1949269298339329951', '1917077971946437171',
                    '1887457570991154046', '1905137082780921471', '1938538127548544502', '1911353123625061642',
                    '1053553576894953602', '1930954268148550608', '1888978417351007848', '1907650482735947096',
                    '1920737431143159646', '1927510059333961808', '1354774767567214920', '1924326314799828901',
                    '1949326757844971366', '1330824435569113217', '1935459250428846222', '1931843907962138916',
                    '1937946240949160550', '1317019749507375805', '1915350837661918950', '1923816141277703835',
                    '1392753179854575276', '1355488927498321743', '1451391803407323272', '1918220282267844500',
                    '1938484909942642948', '1395549062568554707', '1354226522760004385', '1155197472349092012',
                    '1935996001921206273', '1887438535712868482', '1923286738568904528', '1420678832104855328',
                    '1950106607272002684', '1919710167046558060', '1243874179301606364', '1438935513166957858',
                    '1917876881673625908', '1345660260224681112', '1890599961072256156']

EXT_PATH_ORIGIN = "/Users/myeonghyeonjeong/Desktop/DATA_EXT/모아이스/골프자세추출/"

import datetime

now = datetime.datetime.now()
year = str(now.year)
month = str(now.month).zfill(2)
day = str(now.day).zfill(2)

##### 2차 시작 #####
try:
    N_CHA = "labeled_data_1793-1802_" + year + month + day + "/"
    FOLDER_CATEGORY = ""
    PATH = EXT_PATH_ORIGIN + N_CHA
    os.mkdir(PATH)
    for category in FOLDER_CATEGORY:
        FOLDER_CATEGORY = category
        os.mkdir(PATH + FOLDER_CATEGORY)
except FileExistsError as e:
    print("##### 기존에 추출한 " + N_CHA + " 데이터를 모두 삭제하고 폴더 새로 만듦 #####")
    shutil.rmtree(PATH)
    os.mkdir(PATH)

    for category in FOLDER_CATEGORY:
        FOLDER_CATEGORY = category
        os.mkdir(PATH + FOLDER_CATEGORY)

for task_id in TASK_IDS:
    QUERY = "SELECT TPD.result_json, CFU.store_file_name, CFU.store_file_path FROM  TB_PRJ_DATA AS TPD INNER JOIN TB_MEMBER AS TM ON TM.member_id = TPD.work_user INNER JOIN TB_TMPL_IMG_LABEL_DATA AS LD ON TPD.data_idx  = LD.data_idx INNER JOIN CW_SOURCE AS CS ON CS.source_id = LD.src_idx INNER JOIN CW_FILE_UNPACKED AS CFU ON CFU.file_id = CS.file_id WHERE TPD.prj_idx IN (" + str(
        task_id) + ") AND (TPD.prog_state_cd='ALL_FINISHED' OR TPD.prog_state_cd='CHECK_END') ORDER BY store_file_name"
    RESULT = getDatabaseData(QUERY, "prd", "mhjeong", "cworks@34")

    try:
        FOLDER_NAME = RESULT[0][2].split('/')[6]

        if FOLDER_NAME in FOLDER_FRONT:
            EXT_PATH = EXT_PATH_ORIGIN + N_CHA + FOLDER_CATEGORY[0] + FOLDER_NAME + "/"
        if FOLDER_NAME in FOLDER_SIDE:
            EXT_PATH = EXT_PATH_ORIGIN + N_CHA + FOLDER_CATEGORY[1] + FOLDER_NAME + "/"
        if FOLDER_NAME in FOLDER_ETC:
            EXT_PATH = EXT_PATH_ORIGIN + N_CHA + FOLDER_CATEGORY[2] + FOLDER_NAME + "/"

        try:
            os.mkdir(EXT_PATH)
        except FileExistsError:
            print("나면 안되는 오류인데 났음")
            break

        first_file = os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[0]

        if os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[0] == ".DS_Store":
            first_file = os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[1]

        # PIL 라이브러리에 Image 사용해 사이즈 가져오기
        image_path = "/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME + "/" + first_file
        image = Image.open(image_path)
        width, height = image.size

        # 2. 추출한 값들로 엑셀데이터 만들기
        row_idx = 1
        write_wb = Workbook()
        sheet1 = write_wb.active
        xlsx_file_name = EXT_PATH + FOLDER_NAME + ".xlsx"

        sheet1.cell(row=row_idx, column=1).value = "filename"
        sheet1.cell(row=row_idx, column=2).value = "width"
        sheet1.cell(row=row_idx, column=3).value = "height"
        sheet1.cell(row=row_idx, column=4).value = "class"
        sheet1.cell(row=row_idx, column=5).value = "x"
        sheet1.cell(row=row_idx, column=6).value = "y"

        for data in RESULT:
            result_json = json.loads(data[0])
            file_name = data[1]

            for category in result_json:
                for obj in result_json[category]['data']:
                    row_idx = row_idx + 1
                    sheet1.cell(row=row_idx, column=1).value = file_name
                    sheet1.cell(row=row_idx, column=2).value = width
                    sheet1.cell(row=row_idx, column=3).value = height
                    sheet1.cell(row=row_idx, column=4).value = obj['class']
                    sheet1.cell(row=row_idx, column=5).value = obj['x']
                    sheet1.cell(row=row_idx, column=6).value = obj['y']

        print(xlsx_file_name)
        write_wb.save(filename=xlsx_file_name)
        # / 2. 추출한 값들로 엑셀데이터 만들기

        # 3. csv 변환 후 기존 엑셀데이터 삭제
        data_xls = pd.read_excel(xlsx_file_name, 'Sheet', index_col=None)
        data_xls.to_csv(EXT_PATH + FOLDER_NAME + '.csv', encoding='utf-8')
        os.remove(xlsx_file_name)
        # / 3. csv 변환 후 기존 엑셀데이터 삭제
    except IndexError:
        print("IndexError")
        pass
##### 3차 끝 #####

# select U.store_file_path
#
# from  TB_PRJ_DATA     D left join TB_MEMBER M on M.member_id = D.work_user
#
# left join TB_TMPL_IMG_LABEL_DATA L on D.data_idx  = L.data_idx
# left join CW_SOURCE S on S.source_id = L.src_idx
# left join CW_FILE_UNPACKED U  on U.file_id = S.file_id
#
#
# where D.prj_idx in () and D.prog_state_cd="ALL_FINISHED";
