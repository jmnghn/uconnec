"""
[모아이스] 골프자세추출

# 7차 프로젝트아이디: 1921, 1922, 1923, 1924, 1925, 1926, 1928
# 7차 태스크아이디 (83개): 3464,3465,3466,3467,3468,3469,3470,3471,3472,3473,3474,3475,3476,3477,3478,3479,3480,3481,3482,3483,3484,3485,3486,3487,3488,3489,3490,3491,3492,3493,3494,3495,3496,3497,3498,3499,3500,3501,3502,3503,3504,3505,3506,3507,3508,3509,3510,3511,3512,3513,3514,3515,3516,3517,3518,3519,3520,3521,3522,3523,3524,3525,3526,3527,3528,3529,3530,3531,3532,3533,3534,3535,3548,3549,3550,3551,3552,3553,3554,3555,3556,3557,3558
"""

from crowdworks.commonUtil import *

import os
import json
import shutil

from openpyxl import Workbook

from PIL import Image  # 이미지 사이즈 가져오기용 라이브러리
import pandas as pd  # csv 변환용 라이브러리

# 7차 프로젝트아이디: 1921, 1922, 1923, 1924, 1925, 1926, 1928
TASK_IDS = [3464, 3465, 3466, 3467, 3468, 3469, 3470, 3471, 3472, 3473, 3474, 3475, 3476, 3477, 3478, 3479, 3480, 3481,
            3482, 3483, 3484, 3485, 3486, 3487, 3488, 3489, 3490, 3491, 3492, 3493, 3494, 3495, 3496, 3497, 3498, 3499,
            3500, 3501, 3502, 3503, 3504, 3505, 3506, 3507, 3508, 3509, 3510, 3511, 3512, 3513, 3514, 3515, 3516, 3517,
            3518, 3519, 3520, 3521, 3522, 3523, 3524, 3525, 3526, 3527, 3528, 3529, 3530, 3531, 3532, 3533, 3534, 3535,
            3548, 3549, 3550, 3551, 3552, 3553, 3554, 3555, 3556, 3557, 3558]
FOLDER_CATEGORY = ["Front/", "Side/"]  # Front/ - 58 , Side/ - 80, Total - 138
FOLDER_FRONT = ['1810102577073882469', '1815148424669790245', '1812987972878924851', '1825396149629862872',
                '1815703701914935552', '1824093534685578261', '1819263732963076074', '1812422356757583413',
                '1823244110197069477', '1814908975771252396', '1808845660104788659', '1819334253843241227',
                '1807795140565035395', '1809409090339540251', '1813772224872954016', '1825984261355201814',
                '1807243389889748080', '1814332739505090103', '1812150468425333662', '1808026934765879208',
                '1807773871258297146', '1823769230161331670']
FOLDER_SIDE = ['1987367945885531120', '1987245453436375961', '1987550504805720900', '1987253880899304787',
               '1987361291069867996', '1989884240096285401', '1987204953705725170', '1987249122218253382',
               '1987248334536331354', '1989122570921868193', '1989763214107280448', '1987267368272093339',
               '1988474893545900173', '1988404027196572205', '1988118822755161032', '1989236837537011142',
               '1988955420181704370', '1987289421218006082', '1987301795882767253', '1989089704037817930',
               '1989133737753728459', '1987633103754312671', '1987218214385194757', '1988156111241746067',
               '1987858853954608422', '1990352955547895910', '1987254085103440532', '1988546940783492210',
               '1987834614802241161', '1987233977779563996', '1987255272703422805', '1987352550516614935',
               '1987347571659857795', '1990025441173142847', '1987349792595364439', '1987605010079131750',
               '1987254185524210978', '1990007476113036893', '1987317270632185545', '1987668189442922363',
               '1989227217842070359', '1989636102000120643', '1987339057583349684', '1987370461059214234',
               '1987325580277930620', '1987671205323615854', '1990001630250201060', '1987323554774997643',
               '1987288599218334765', '1987314072105801222', '1988389669189396609', '1987363207153746499',
               '1988158447083097515', '1987415484421496360', '1989210180537075089', '1987353421423275854',
               '1988247408103295969', '1989698778240918497', '1987366151546861790', '1988273948367153155',
               '1987322008208447315']

EXT_PATH_ORIGIN = "/Users/myeonghyeonjeong/Desktop/DATA_EXT/모아이스/골프자세추출/"

import datetime

now = datetime.datetime.now()
year = str(now.year)
month = str(now.month).zfill(2)
day = str(now.day).zfill(2)

##### 7차 시작 #####
try:
    N_CHA = "labeled_data_1921-1928_" + year + month + day + "/"
    FOLDER_CATEGORY = ""
    PATH = EXT_PATH_ORIGIN + N_CHA
    os.mkdir(PATH)
    for category in FOLDER_CATEGORY:
        FOLDER_CATEGORY = category
        os.mkdir(PATH + FOLDER_CATEGORY)
except FileExistsError as e:
    print("##### 기존에 추출한 " + N_CHA + " 데이터를 모두 삭제하고 폴더 새로 만듦 #####")
    shutil.rmtree(PATH)
    os.mkdir(PATH)

    for category in FOLDER_CATEGORY:
        FOLDER_CATEGORY = category
        os.mkdir(PATH + FOLDER_CATEGORY)

for task_id in TASK_IDS:
    QUERY = "SELECT TPD.result_json, CFU.store_file_name, CFU.store_file_path FROM  TB_PRJ_DATA AS TPD INNER JOIN TB_MEMBER AS TM ON TM.member_id = TPD.work_user INNER JOIN TB_TMPL_IMG_LABEL_DATA AS LD ON TPD.data_idx  = LD.data_idx INNER JOIN CW_SOURCE AS CS ON CS.source_id = LD.src_idx INNER JOIN CW_FILE_UNPACKED AS CFU ON CFU.file_id = CS.file_id WHERE TPD.prj_idx IN (" + str(
        task_id) + ") AND (TPD.prog_state_cd='ALL_FINISHED' OR TPD.prog_state_cd='CHECK_END') ORDER BY store_file_name"
    RESULT = getDatabaseData(QUERY, "prd", "mhjeong", "cworks@34")

    try:
        FOLDER_NAME = RESULT[0][2].split('/')[6]

        if FOLDER_NAME in FOLDER_FRONT:
            EXT_PATH = EXT_PATH_ORIGIN + N_CHA + FOLDER_CATEGORY[0] + FOLDER_NAME + "/"
        if FOLDER_NAME in FOLDER_SIDE:
            EXT_PATH = EXT_PATH_ORIGIN + N_CHA + FOLDER_CATEGORY[1] + FOLDER_NAME + "/"

        try:
            os.mkdir(EXT_PATH)
        except FileExistsError:
            print("나면 안되는 오류인데 났음")
            break

        first_file = os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[0]

        if os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[0] == ".DS_Store":
            first_file = os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[1]

        # PIL 라이브러리에 Image 사용해 사이즈 가져오기
        image_path = "/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME + "/" + first_file
        image = Image.open(image_path)
        width, height = image.size

        # 2. 추출한 값들로 엑셀데이터 만들기
        row_idx = 1
        write_wb = Workbook()
        sheet1 = write_wb.active
        xlsx_file_name = EXT_PATH + FOLDER_NAME + ".xlsx"

        sheet1.cell(row=row_idx, column=1).value = "filename"
        sheet1.cell(row=row_idx, column=2).value = "width"
        sheet1.cell(row=row_idx, column=3).value = "height"
        sheet1.cell(row=row_idx, column=4).value = "class"
        sheet1.cell(row=row_idx, column=5).value = "x"
        sheet1.cell(row=row_idx, column=6).value = "y"

        for data in RESULT:
            result_json = json.loads(data[0])
            file_name = data[1]

            for category in result_json:
                for obj in result_json[category]['data']:
                    row_idx = row_idx + 1
                    sheet1.cell(row=row_idx, column=1).value = file_name
                    sheet1.cell(row=row_idx, column=2).value = width
                    sheet1.cell(row=row_idx, column=3).value = height
                    sheet1.cell(row=row_idx, column=4).value = obj['class']
                    sheet1.cell(row=row_idx, column=5).value = obj['x']
                    sheet1.cell(row=row_idx, column=6).value = obj['y']

        print(xlsx_file_name)
        write_wb.save(filename=xlsx_file_name)
        # / 2. 추출한 값들로 엑셀데이터 만들기

        # 3. csv 변환 후 기존 엑셀데이터 삭제
        data_xls = pd.read_excel(xlsx_file_name, 'Sheet', index_col=None)
        data_xls.to_csv(EXT_PATH + FOLDER_NAME + '.csv', encoding='utf-8')
        os.remove(xlsx_file_name)
        # / 3. csv 변환 후 기존 엑셀데이터 삭제
    except IndexError:
        print("IndexError")
        pass
##### 7차 끝 #####

# select U.store_file_path
#
# from  TB_PRJ_DATA     D left join TB_MEMBER M on M.member_id = D.work_user
#
# left join TB_TMPL_IMG_LABEL_DATA L on D.data_idx  = L.data_idx
# left join CW_SOURCE S on S.source_id = L.src_idx
# left join CW_FILE_UNPACKED U  on U.file_id = S.file_id
#
#
# where D.prj_idx in () and D.prog_state_cd="ALL_FINISHED";
