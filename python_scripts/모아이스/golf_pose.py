from crowdworks.commonUtil import *

import os
import shutil
import json
import uuid
from openpyxl import Workbook

from PIL import Image  # 이미지 사이즈 가져오기용 라이브러리
import pandas as pd  # csv 변환용 라이브러리

# 2 차 project_id (1717, 1718, 1719, 1720, 1721, 1722, 1723, 1724, 1725, 1726, 1727, 1728, 1729, 1730, 1731, 1732, 1733, 1734, 1735, 1736, 1737, 1738, 1739, 1740, 1741)
# 3 차 project_id (1793, 1794, 1795, 1796, 1797, 1798, 1799, 1800, 1801, 1802)
task_id_list = [2413, 2414, 2415, 2416, 2417, 2418, 2419, 2420, 2421, 2422, 2423, 2424, 2425, 2426, 2427, 2428, 2429,
                2430, 2431, 2432, 2433, 2434, 2435, 2436, 2437, 2438, 2439, 2440, 2441, 2442, 2443, 2444, 2445, 2446,
                2447, 2448, 2449, 2450, 2451, 2452, 2453, 2454, 2455, 2456, 2457, 2458, 2459, 2460, 2461, 2462, 2463,
                2464, 2465, 2466, 2467, 2468, 2469, 2470, 2471, 2472, 2473, 2474, 2475, 2476, 2477, 2478, 2479, 2480,
                2481, 2482, 2483, 2484, 2485, 2486, 2487, 2488, 2489, 2490, 2491, 2492, 2493, 2494, 2495, 2496, 2497,
                2498, 2499, 2500, 2501, 2502, 2503, 2504, 2505, 2506, 2507, 2508, 2509, 2510, 2511, 2512, 2513, 2514,
                2515, 2516, 2517, 2518, 2519, 2520, 2521, 2522, 2523, 2524, 2525, 2526, 2527, 2528, 2529, 2530, 2531,
                2532, 2533, 2534, 2535, 2536, 2537, 2538, 2539, 2540, 2541, 2542, 2543, 2544, 2545, 2546, 2547, 2548,
                2549, 2550, 2551, 2552, 2553, 2554, 2555, 2556, 2557, 2558, 2559, 2560, 2561, 2562, 2563, 2564, 2565,
                2566, 2567, 2568, 2569, 2570, 2571, 2572, 2573, 2574, 2575, 2576, 2577, 2578, 2579, 2580, 2581, 2582,
                2583, 2584, 2585, 2586, 2587, 2588, 2589, 2590, 2591, 2592, 2593, 2594, 2595, 2596, 2597, 2598, 2599,
                2600, 2601, 2602, 2603, 2604, 2605, 2606, 2607, 2608, 2609, 2610, 2611, 2612, 2613, 2614, 2615, 2616,
                2617, 2618, 2619, 2620, 2621, 2622, 2623, 2624, 2625, 2626, 2627, 2628, 2629, 2630, 2631, 2632, 2633,
                2634, 2635, 2636, 2637, 2638, 2639, 2640, 2641, 2642, 2643, 2644, 2645, 2646, 2647, 2648, 2649, 2650,
                2651, 2652, 2653, 2654, 2655, 2656, 2657, 2658, 2659, 2660, 2661, 2662, 2663, 2664, 2665, 2666, 2667,
                2668, 2669, 2670, 2671, 2672, 2673, 2674, 2675, 2676, 2677, 2678, 2679, 2680, 2681, 2682, 2683, 2684,
                2685, 2686, 2687, 2688, 2689, 2690, 2691, 2692, 2693, 2694, 2695, 2696, 2697, 2698, 2699, 2700, 2701,
                2702, 2703, 2704, 2705, 2706, 2707, 2708, 2709, 2710, 2711, 2712, 2713, 2714, 2715, 2716, 2717, 2718,
                2719, 2720, 2721, 2722, 2723, 2724, 2725, 2726, 2727, 2728, 2729, 2730, 2731, 2732, 2733, 2734, 2735,
                2736, 2737, 2738, 2739, 2740, 2741, 2742, 2743, 2744, 2808, 2809, 2810, 2811, 2812, 2813, 2814, 2815,
                2816, 2817, 2818, 2819, 2820, 2821, 2822, 2823, 2824, 2825, 2826, 2827, 2828, 2829, 2830, 2831, 2832,
                2833, 2834, 2835, 2836, 2837, 2838, 2839, 2840, 2841, 2842, 2843, 2844, 2845, 2846, 2847, 2848, 2849,
                2850, 2851, 2852, 2853, 2854, 2855, 2856, 2857, 2858, 2859, 2860, 2861, 2862, 2863, 2864, 2865, 2866,
                2867, 2868, 2869, 2870, 2871, 2872, 2873, 2874, 2875, 2876, 2877, 2878, 2879, 2880, 2881, 2882, 2883,
                2884, 2885, 2886, 2887, 2888, 2889, 2890, 2891, 2892, 2893, 2894, 2895, 2896, 2897, 2898, 2899, 2900,
                2901, 2902, 2903, 2904, 2905, 2906, 2907]

# 4 차 project_id (1838, 1839, 1840, 1841, 1842, 1843)
# task_id_list = [3030, 3031, 3032, 3033, 3034, 3035, 3036, 3037, 3038, 3039, 3040, 3041, 3042, 3043, 3044, 3045, 3046,
#                 3047, 3048, 3049, 3050, 3051, 3052, 3053, 3054, 3055, 3056, 3057, 3058, 3059, 3060, 3061, 3062, 3063,
#                 3064, 3065, 3066, 3067, 3068, 3069, 3070, 3071, 3072, 3073, 3074, 3075, 3076, 3077, 3078, 3079, 3080,
#                 3081, 3082, 3083, 3084, 3085, 3086, 3087, 3088, 3089, 3090, 3091, 3092, 3093, 3094, 3095, 3096, 3097,
#                 3098, 3099, 3100, 3101, 3102, 3103, 3104, 3105, 3106, 3107, 3108, 3109, 3110]

complete_task = []
complete_task2 = []
empty_task = []

print('task 갯수: ', len(task_id_list))

for task_id in task_id_list:
    sql_result = "select D.result_json, U.store_file_name, U.store_file_path " \
                 "from  TB_PRJ_DATA     D inner join TB_MEMBER M on M.member_id = D.work_user " \
                 "inner join TB_TMPL_IMG_LABEL_DATA L on D.data_idx  = L.data_idx " \
                 "inner join CW_SOURCE S on S.source_id = L.src_idx " \
                 "inner join CW_FILE_UNPACKED U  on U.file_id = S.file_id " \
                 "where D.prj_idx in (" + str(
        task_id) + ") and (D.prog_state_cd='ALL_FINISHED' OR D.prog_state_cd='CHECK_END') ORDER BY store_file_name"

    data_list = getDatabaseData(sql_result, "", "", "")

    try:
        # print("source_name: ", data_list[0][2].split('/')[6])
        print(task_id)
    except IndexError:
        print(task_id)
        print(data_list)

    try:
        source_name = data_list[0][2].split('/')[6]
        complete_item = {
            "task_id": 0,
            "source_name": 0,
            "file_length": 0
        }

        # 1. 추출 파일 저장할 디렉토리 생성 (소스의 압축파일 이름)
        output_path = "/Users/myeonghyeonjeong/Desktop/GolfPose/" + source_name
        try:
            if not (os.path.isdir(output_path)):
                os.makedirs(os.path.join(output_path))
                # print(output_path + " 폴더 생성 성공")
        except OSError as e:
            if e.errno != e.errno.EEXIST:
                print(output_path + " 폴더 생성 실패")
                raise
        # / 1. 추출 파일 저장할 디렉토리 생성 (소스의 압축파일 이름)

        # print(os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + source_name)[0])
        # break

        first_file = os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + source_name)[0]

        if os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + source_name)[0] == ".DS_Store":
            first_file = os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + source_name)[1]

        # PIL 라이브러리에 Image 사용해 사이즈 가져오기
        image_path = "/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + source_name + "/" + first_file
        image = Image.open(image_path)
        width, height = image.size

        # 2. 추출한 값들로 엑셀데이터 만들기
        row_idx = 1
        write_wb = Workbook()
        sheet1 = write_wb.active
        xlsx_file_name = output_path + "/" + source_name + ".xlsx"

        sheet1.cell(row=row_idx, column=1).value = "filename"
        sheet1.cell(row=row_idx, column=2).value = "width"
        sheet1.cell(row=row_idx, column=3).value = "height"
        sheet1.cell(row=row_idx, column=4).value = "class"
        sheet1.cell(row=row_idx, column=5).value = "x"
        sheet1.cell(row=row_idx, column=6).value = "y"

        for data in data_list:
            result_json = json.loads(data[0])
            file_name = data[1]

            print(result_json)

            for category in result_json:
                for obj in result_json[category]['data']:
                    row_idx = row_idx + 1
                    sheet1.cell(row=row_idx, column=1).value = file_name
                    sheet1.cell(row=row_idx, column=2).value = width
                    sheet1.cell(row=row_idx, column=3).value = height
                    sheet1.cell(row=row_idx, column=4).value = obj['class']
                    sheet1.cell(row=row_idx, column=5).value = obj['x']
                    sheet1.cell(row=row_idx, column=6).value = obj['y']

        print(xlsx_file_name)
        write_wb.save(filename=xlsx_file_name)
        # / 2. 추출한 값들로 엑셀데이터 만들기

        # 3. csv 변환 후 기존 엑셀데이터 삭제
        data_xls = pd.read_excel(xlsx_file_name, 'Sheet', index_col=None)
        data_xls.to_csv(output_path + "/" + source_name + '.csv', encoding='utf-8')
        os.remove(xlsx_file_name)
        # / 3. csv 변환 후 기존 엑셀데이터 삭제
        complete_task.append(task_id)
        # print(task_id, " task 의 ", source_name, " 의 파일 갯수: ", len(data_list))
        complete_item['task_id'] = task_id
        complete_item['source_name'] = source_name
        complete_item['file_length'] = len(data_list)
        complete_task2.append(complete_item)
    except IndexError:
        # print("ALL_FINISHED 데이터가 없는 TASK ID: ", task_id)
        empty_task.append(task_id)

print('##### complete task list ##### \n', complete_task)
print('##### complete task2 list ##### \n', complete_task2)
print('##### empty task list ##### \n', empty_task)

# select U.store_file_path
#
# from  TB_PRJ_DATA     D left join TB_MEMBER M on M.member_id = D.work_user
#
# left join TB_TMPL_IMG_LABEL_DATA L on D.data_idx  = L.data_idx
# left join CW_SOURCE S on S.source_id = L.src_idx
# left join CW_FILE_UNPACKED U  on U.file_id = S.file_id
#
#
# where D.prj_idx in () and D.prog_state_cd="ALL_FINISHED";
