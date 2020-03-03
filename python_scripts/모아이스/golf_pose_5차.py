"""
[모아이스] 골프자세추출

# 5차 프로젝트아이디: 1851,1852,1853,1854,1855,1856
# 5차 태스크아이디 (78개): 3123,3124,3125,3126,3127,3128,3129,3130,3131,3132,3133,3134,3135,3136,3137,3138,3139,3140,3141,3142,3143,3144,3145,3146,3147,3148,3149,3150,3151,3152,3153,3154,3155,3156,3157,3158,3159,3160,3161,3162,3163,3164,3165,3166,3167,3168,3169,3170,3171,3172,3173,3174,3175,3176,3177,3178,3179,3180,3181,3182,3183,3184,3185,3186,3187,3188,3189,3190,3191,3192,3193,3194,3195,3196,3197,3198,3199,3200
"""

from crowdworks.commonUtil import *

import os
import json
import shutil

from openpyxl import Workbook

from PIL import Image  # 이미지 사이즈 가져오기용 라이브러리
import pandas as pd  # csv 변환용 라이브러리

# 5차 프로젝트아이디: 1851,1852,1853,1854,1855,1856
TASK_IDS = [3123, 3124, 3125, 3126, 3127, 3128, 3129, 3130, 3131, 3132, 3133, 3134, 3135, 3136, 3137, 3138, 3139, 3140,
            3141, 3142, 3143, 3144, 3145, 3146, 3147, 3148, 3149, 3150, 3151, 3152, 3153, 3154, 3155, 3156, 3157, 3158,
            3159, 3160, 3161, 3162, 3163, 3164, 3165, 3166, 3167, 3168, 3169, 3170, 3171, 3172, 3173, 3174, 3175, 3176,
            3177, 3178, 3179, 3180, 3181, 3182, 3183, 3184, 3185, 3186, 3187, 3188, 3189, 3190, 3191, 3192, 3193, 3194,
            3195, 3196, 3197, 3198, 3199, 3200]
FOLDER_CATEGORY = ["Front/", "Side/"]  # Front/ - 24 , Side/ - 54, Total - 78
FOLDER_FRONT = ['1556601442886524514', '1550526435796907874', '1557215483605810044', '1552137801175171716',
                '1546330438489722300', '1566027329116110047', '1551418459693724488', '1547720810151283398',
                '1541990635509181096', '1537639728604715892', '1569676901256988000', '1554934890473932933',
                '1552752165872658233', '1559317654425963427', '1538331791105849942', '1536076021425435773',
                '1549759826832604059', '1562141866551298442', '1545459220999337172', '1546395539818298977',
                '1560225288254962544', '1556295728824621892', '1542646341816363241', '1550537985660646441']
FOLDER_SIDE = ['1973725062843691266', '1974387405604791732', '1966465909851001963', '1971786962776289699',
               '1971703681239053706', '1973029597084987036', '1971780067568610818', '1972928675050750516',
               '1965729146178187030', '1968143439200859371', '1964373471770491835', '1964495038487771136',
               '1964438788182514924', '1974100973724145684', '1965067186733156487', '1973406094077697874',
               '1974477769718922189', '1971857281886623746', '1971780254909766722', '1972632696976299779',
               '1965535712695854304', '1968412941830348663', '1964676042610277023', '1970167140440320253',
               '1968413796292921286', '1970216658871402712', '1975210779717445564', '1968824401073177103',
               '1968221630564329117', '1974811078140179336', '1972430701468106610', '1970048966973525373',
               '1973738520595311266', '1969721749043283061', '1972372279780088264', '1974836222280157335',
               '1972639550293363014', '1973075076015590198', '1968824491795961261', '1971717604877068918',
               '1969386700951495906', '1973906867592937715', '1967045963924587900', '1969363403688548906',
               '1969427685400085956', '1968062523224017449', '1967587613390170802', '1968833997682722519',
               '1975223123872942610', '1972597894427896625', '1964533340914022378', '1964673910259765134',
               '1970999248687470537', '1970344079402699748']

EXT_PATH_ORIGIN = "/Users/myeonghyeonjeong/Desktop/DATA_EXT/모아이스/골프자세추출/"

import datetime

now = datetime.datetime.now()
year = str(now.year)
month = str(now.month).zfill(2)
day = str(now.day).zfill(2)

##### 5차 시작 #####
try:
    N_CHA = "labeled_data_1851-1856_" + year + month + day + "/"
    PATH = EXT_PATH_ORIGIN + N_CHA
    os.mkdir(PATH)
    for category in FOLDER_CATEGORY:
        os.mkdir(PATH + category)
except FileExistsError as e:
    print("##### 기존에 추출한 " + N_CHA + " 데이터를 모두 삭제하고 폴더 새로 만듦 #####")
    shutil.rmtree(PATH)
    os.mkdir(PATH)

    for category in FOLDER_CATEGORY:
        os.mkdir(PATH + category)

for task_id in TASK_IDS:
    QUERY = "SELECT TPD.result_json, CFU.store_file_name, CFU.store_file_path FROM  TB_PRJ_DATA AS TPD INNER JOIN TB_MEMBER AS TM ON TM.member_id = TPD.work_user INNER JOIN TB_TMPL_IMG_LABEL_DATA AS LD ON TPD.data_idx  = LD.data_idx INNER JOIN CW_SOURCE AS CS ON CS.source_id = LD.src_idx INNER JOIN CW_FILE_UNPACKED AS CFU ON CFU.file_id = CS.file_id WHERE TPD.prj_idx IN (" + str(
        task_id) + ") AND (TPD.prog_state_cd='ALL_FINISHED' OR TPD.prog_state_cd='CHECK_END') ORDER BY store_file_name"
    RESULT = getDatabaseData(QUERY, "", "", "")

    try:
        FOLDER_NAME = RESULT[0][2].split('/')[6]

        if FOLDER_NAME in FOLDER_FRONT:
            EXT_PATH = EXT_PATH_ORIGIN + N_CHA + FOLDER_CATEGORY[0] + FOLDER_NAME + "/"
        if FOLDER_NAME in FOLDER_SIDE:
            EXT_PATH = EXT_PATH_ORIGIN + N_CHA + FOLDER_CATEGORY[1] + FOLDER_NAME + "/"

        try:
            os.mkdir(EXT_PATH)
        except FileExistsError:
            print("나면 안되는 오류인데 났음")
            break

        first_file = os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[0]

        if os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[0] == ".DS_Store":
            first_file = os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[1]

        # PIL 라이브러리에 Image 사용해 사이즈 가져오기
        image_path = "/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME + "/" + first_file
        image = Image.open(image_path)
        width, height = image.size

        # 2. 추출한 값들로 엑셀데이터 만들기
        row_idx = 1
        write_wb = Workbook()
        sheet1 = write_wb.active
        xlsx_file_name = EXT_PATH + FOLDER_NAME + ".xlsx"

        sheet1.cell(row=row_idx, column=1).value = "filename"
        sheet1.cell(row=row_idx, column=2).value = "width"
        sheet1.cell(row=row_idx, column=3).value = "height"
        sheet1.cell(row=row_idx, column=4).value = "class"
        sheet1.cell(row=row_idx, column=5).value = "x"
        sheet1.cell(row=row_idx, column=6).value = "y"

        for data in RESULT:
            result_json = json.loads(data[0])
            file_name = data[1]

            for category in result_json:
                for obj in result_json[category]['data']:
                    row_idx = row_idx + 1
                    sheet1.cell(row=row_idx, column=1).value = file_name
                    sheet1.cell(row=row_idx, column=2).value = width
                    sheet1.cell(row=row_idx, column=3).value = height
                    sheet1.cell(row=row_idx, column=4).value = obj['class']
                    sheet1.cell(row=row_idx, column=5).value = obj['x']
                    sheet1.cell(row=row_idx, column=6).value = obj['y']

        print(xlsx_file_name)
        write_wb.save(filename=xlsx_file_name)
        # / 2. 추출한 값들로 엑셀데이터 만들기

        # 3. csv 변환 후 기존 엑셀데이터 삭제
        data_xls = pd.read_excel(xlsx_file_name, 'Sheet', index_col=None)
        data_xls.to_csv(EXT_PATH + FOLDER_NAME + '.csv', encoding='utf-8')
        os.remove(xlsx_file_name)
        # / 3. csv 변환 후 기존 엑셀데이터 삭제
    except IndexError:
        print("IndexError")
        pass
##### 5차 끝 #####

# select U.store_file_path
#
# from  TB_PRJ_DATA     D left join TB_MEMBER M on M.member_id = D.work_user
#
# left join TB_TMPL_IMG_LABEL_DATA L on D.data_idx  = L.data_idx
# left join CW_SOURCE S on S.source_id = L.src_idx
# left join CW_FILE_UNPACKED U  on U.file_id = S.file_id
#
#
# where D.prj_idx in () and D.prog_state_cd="ALL_FINISHED";
