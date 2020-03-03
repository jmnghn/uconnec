"""
[모아이스] 골프자세추출

# 8차 프로젝트아이디:
# 8차 태스크아이디 (239개):
"""

from crowdworks.commonUtil import *

import os
import json
import shutil

from openpyxl import Workbook

from PIL import Image  # 이미지 사이즈 가져오기용 라이브러리
import pandas as pd  # csv 변환용 라이브러리

# 8차 프로젝트아이디:
TASK_IDS = []
FOLDER_CATEGORY = ["Front/", "Side/"]  # Front/ - 58 , Side/ - 80, Total - 138
FOLDER_FRONT = []
FOLDER_SIDE = []

EXT_PATH_ORIGIN = "/Users/myeonghyeonjeong/Desktop/DATA_EXT/모아이스/골프자세추출/"

import datetime

now = datetime.datetime.now()
year = str(now.year)
month = str(now.month).zfill(2)
day = str(now.day).zfill(2)

##### 8차 시작 #####
try:
    # TODO: 
    N_CHA = "labeled_data_1921-1928_" + year + month + day + "/"
    FOLDER_CATEGORY = ""
    PATH = EXT_PATH_ORIGIN + N_CHA
    os.mkdir(PATH)
    for category in FOLDER_CATEGORY:
        FOLDER_CATEGORY = category
        os.mkdir(PATH + FOLDER_CATEGORY)
except FileExistsError as e:
    print("##### 기존에 추출한 " + N_CHA + " 데이터를 모두 삭제하고 폴더 새로 만듦 #####")
    shutil.rmtree(PATH)
    os.mkdir(PATH)

    for category in FOLDER_CATEGORY:
        FOLDER_CATEGORY = category
        os.mkdir(PATH + FOLDER_CATEGORY)

for task_id in TASK_IDS:
    QUERY = "SELECT TPD.result_json, CFU.store_file_name, CFU.store_file_path FROM  TB_PRJ_DATA AS TPD INNER JOIN TB_MEMBER AS TM ON TM.member_id = TPD.work_user INNER JOIN TB_TMPL_IMG_LABEL_DATA AS LD ON TPD.data_idx  = LD.data_idx INNER JOIN CW_SOURCE AS CS ON CS.source_id = LD.src_idx INNER JOIN CW_FILE_UNPACKED AS CFU ON CFU.file_id = CS.file_id WHERE TPD.prj_idx IN (" + str(
        task_id) + ") AND (TPD.prog_state_cd='ALL_FINISHED' OR TPD.prog_state_cd='CHECK_END') ORDER BY store_file_name"
    RESULT = getDatabaseData(QUERY, "", "", "")

    try:
        FOLDER_NAME = RESULT[0][2].split('/')[6]

        if FOLDER_NAME in FOLDER_FRONT:
            EXT_PATH = EXT_PATH_ORIGIN + N_CHA + FOLDER_CATEGORY[0] + FOLDER_NAME + "/"
        if FOLDER_NAME in FOLDER_SIDE:
            EXT_PATH = EXT_PATH_ORIGIN + N_CHA + FOLDER_CATEGORY[1] + FOLDER_NAME + "/"

        try:
            os.mkdir(EXT_PATH)
        except FileExistsError:
            print("나면 안되는 오류인데 났음")
            break

        first_file = os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[0]

        if os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[0] == ".DS_Store":
            first_file = os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[1]

        # PIL 라이브러리에 Image 사용해 사이즈 가져오기
        image_path = "/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME + "/" + first_file
        image = Image.open(image_path)
        width, height = image.size

        # 2. 추출한 값들로 엑셀데이터 만들기
        row_idx = 1
        write_wb = Workbook()
        sheet1 = write_wb.active
        xlsx_file_name = EXT_PATH + FOLDER_NAME + ".xlsx"

        sheet1.cell(row=row_idx, column=1).value = "filename"
        sheet1.cell(row=row_idx, column=2).value = "width"
        sheet1.cell(row=row_idx, column=3).value = "height"
        sheet1.cell(row=row_idx, column=4).value = "class"
        sheet1.cell(row=row_idx, column=5).value = "x"
        sheet1.cell(row=row_idx, column=6).value = "y"

        for data in RESULT:
            result_json = json.loads(data[0])
            file_name = data[1]

            for category in result_json:
                for obj in result_json[category]['data']:
                    row_idx = row_idx + 1
                    sheet1.cell(row=row_idx, column=1).value = file_name
                    sheet1.cell(row=row_idx, column=2).value = width
                    sheet1.cell(row=row_idx, column=3).value = height
                    sheet1.cell(row=row_idx, column=4).value = obj['class']
                    sheet1.cell(row=row_idx, column=5).value = obj['x']
                    sheet1.cell(row=row_idx, column=6).value = obj['y']

        print(xlsx_file_name)
        write_wb.save(filename=xlsx_file_name)
        # / 2. 추출한 값들로 엑셀데이터 만들기

        # 3. csv 변환 후 기존 엑셀데이터 삭제
        data_xls = pd.read_excel(xlsx_file_name, 'Sheet', index_col=None)
        data_xls.to_csv(EXT_PATH + FOLDER_NAME + '.csv', encoding='utf-8')
        os.remove(xlsx_file_name)
        # / 3. csv 변환 후 기존 엑셀데이터 삭제
    except IndexError:
        print("IndexError")
        pass
##### 8차 끝 #####

# select U.store_file_path
#
# from  TB_PRJ_DATA     D left join TB_MEMBER M on M.member_id = D.work_user
#
# left join TB_TMPL_IMG_LABEL_DATA L on D.data_idx  = L.data_idx
# left join CW_SOURCE S on S.source_id = L.src_idx
# left join CW_FILE_UNPACKED U  on U.file_id = S.file_id
#
#
# where D.prj_idx in () and D.prog_state_cd="ALL_FINISHED";
