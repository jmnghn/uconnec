"""
[모아이스] 골프자세추출

# 4차 프로젝트아이디: 1838, 1839, 1840, 1841, 1842, 1843
# 4차 태스크아이디 (81개): 3030,3031,3032,3033,3034,3035,3036,3037,3038,3039,3040,3041,3042,3043,3044,3045,3046,3047,3048,3049,3050,3051,3052,3053,3054,3055,3056,3057,3058,3059,3060,3061,3062,3063,3064,3065,3066,3067,3068,3069,3070,3071,3072,3073,3074,3075,3076,3077,3078,3079,3080,3081,3082,3083,3084,3085,3086,3087,3088,3089,3090,3091,3092,3093,3094,3095,3096,3097,3098,3099,3100,3101,3102,3103,3104,3105,3106,3107,3108,3109,3110
"""

from crowdworks.commonUtil import *

import os
import json
import shutil

from openpyxl import Workbook

from PIL import Image  # 이미지 사이즈 가져오기용 라이브러리
import pandas as pd  # csv 변환용 라이브러리

# 4차 프로젝트아이디: 1838, 1839, 1840, 1841, 1842, 1843
TASK_IDS = [3030, 3031, 3032, 3033, 3034, 3035, 3036, 3037, 3038, 3039, 3040, 3041, 3042, 3043, 3044, 3045, 3046,
            3047, 3048, 3049, 3050, 3051, 3052, 3053, 3054, 3055, 3056, 3057, 3058, 3059, 3060, 3061, 3062, 3063,
            3064, 3065, 3066, 3067, 3068, 3069, 3070, 3071, 3072, 3073, 3074, 3075, 3076, 3077, 3078, 3079, 3080,
            3081, 3082, 3083, 3084, 3085, 3086, 3087, 3088, 3089, 3090, 3091, 3092, 3093, 3094, 3095, 3096, 3097,
            3098, 3099, 3100, 3101, 3102, 3103, 3104, 3105, 3106, 3107, 3108, 3109, 3110]
FOLDER_CATEGORY = ["Front/", "Side/"]  # Front/ - 41 , Side/ - 40, Total - 81
FOLDER_FRONT = ['1770985477445370864', '1492008100528773653', '1784884435289062420', '1778179936620990360',
                '1778943033006531138', '1374554642532728133', '1470035902801039036', '1777612592169911320',
                '1779152702581664102', '1775499690901580505', '1783420682266124138', '1784635376074386962',
                '1775467564798745650', '1787449057838116997', '1772536104311262674', '1774094837382694417',
                '1780467946817074358', '1787123328423804110', '1389549428323147020', '1776706922205876588',
                '1501362305383541257', '1774680835958881277', '1786778120797235876', '1770375059019914888',
                '1782017783492554988', '1776730036997711748', '1782813324438694601', '1782003985197865624',
                '1776820737899565173', '1773201688556199343', '1368326321679353666', '1771747003392320720',
                '1790494999143327584', '1776717903505284145', '1256941823243648355', '1775870565455534374',
                '1786629762752534422', '1774071735809937557', '1777967023834721220', '1776983136845067952',
                '1780572544773098322']
FOLDER_SIDE = ['1962922529366850628', '1963837870105424179', '1959424830098295247', '1963223516263182171',
               '1961086125478967707', '1959262348801025444', '1962083776838525538', '1958107363417544386',
               '1957760782776495764', '1961645754801278608', '1957265374387370887', '1960701636381004799',
               '1960934550369796277', '1963866717035054021', '1963237459984077517', '1957277559092532164',
               '1961474062695231567', '1961501006417572903', '1957270270667754438', '1958846063851669501',
               '1960232841121295389', '1963687071312089788', '1961582438695359364', '1957739633904706407',
               '1962064571733918321', '1958134849916932206', '1957997029179309335', '1959490169108583528',
               '1963635409281268948', '1960705437042472818', '1959565607498328478', '1964283739224469186',
               '1957690894741618508', '1957236473855248355', '1964256938939410193', '1962973682418655635',
               '1961281767915955772', '1963864707065774913', '1961039692662570134', '1958574370008839546']

EXT_PATH_ORIGIN = "/Users/myeonghyeonjeong/Desktop/DATA_EXT/모아이스/골프자세추출/"

import datetime

now = datetime.datetime.now()
year = str(now.year)
month = str(now.month).zfill(2)
day = str(now.day).zfill(2)

##### 4차 시작 #####
try:
    N_CHA = "labeled_data_1838-1843_" + year + month + day + "/"
    PATH = EXT_PATH_ORIGIN + N_CHA
    os.mkdir(PATH)
    for category in FOLDER_CATEGORY:
        os.mkdir(PATH + category)
except FileExistsError as e:
    print("##### 기존에 추출한 " + N_CHA + " 데이터를 모두 삭제하고 폴더 새로 만듦 #####")
    shutil.rmtree(PATH)
    os.mkdir(PATH)

    for category in FOLDER_CATEGORY:
        os.mkdir(PATH + category)

for task_id in TASK_IDS:
    QUERY = "SELECT TPD.result_json, CFU.store_file_name, CFU.store_file_path FROM  TB_PRJ_DATA AS TPD INNER JOIN TB_MEMBER AS TM ON TM.member_id = TPD.work_user INNER JOIN TB_TMPL_IMG_LABEL_DATA AS LD ON TPD.data_idx  = LD.data_idx INNER JOIN CW_SOURCE AS CS ON CS.source_id = LD.src_idx INNER JOIN CW_FILE_UNPACKED AS CFU ON CFU.file_id = CS.file_id WHERE TPD.prj_idx IN (" + str(
        task_id) + ") AND (TPD.prog_state_cd='ALL_FINISHED' OR TPD.prog_state_cd='CHECK_END') ORDER BY store_file_name"
    RESULT = getDatabaseData(QUERY, "", "", "")

    try:
        FOLDER_NAME = RESULT[0][2].split('/')[6]

        if FOLDER_NAME in FOLDER_FRONT:
            EXT_PATH = EXT_PATH_ORIGIN + N_CHA + FOLDER_CATEGORY[0] + FOLDER_NAME + "/"
        if FOLDER_NAME in FOLDER_SIDE:
            EXT_PATH = EXT_PATH_ORIGIN + N_CHA + FOLDER_CATEGORY[1] + FOLDER_NAME + "/"
    except IndexError:
        print("IndexError")
        pass

    try:
        os.mkdir(EXT_PATH)
    except FileExistsError:
        print("나면 안되는 오류인데 났음")
        break

    first_file = os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[0]

    if os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[0] == ".DS_Store":
        first_file = os.listdir("/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME)[1]

    # PIL 라이브러리에 Image 사용해 사이즈 가져오기
    image_path = "/Users/myeonghyeonjeong/Desktop/GolfPose/ImageSize/" + FOLDER_NAME + "/" + first_file
    image = Image.open(image_path)
    width, height = image.size

    # 2. 추출한 값들로 엑셀데이터 만들기
    row_idx = 1
    write_wb = Workbook()
    sheet1 = write_wb.active
    xlsx_file_name = EXT_PATH + FOLDER_NAME + ".xlsx"

    sheet1.cell(row=row_idx, column=1).value = "filename"
    sheet1.cell(row=row_idx, column=2).value = "width"
    sheet1.cell(row=row_idx, column=3).value = "height"
    sheet1.cell(row=row_idx, column=4).value = "class"
    sheet1.cell(row=row_idx, column=5).value = "x"
    sheet1.cell(row=row_idx, column=6).value = "y"

    for data in RESULT:
        result_json = json.loads(data[0])
        file_name = data[1]

        for category in result_json:
            for obj in result_json[category]['data']:
                row_idx = row_idx + 1
                sheet1.cell(row=row_idx, column=1).value = file_name
                sheet1.cell(row=row_idx, column=2).value = width
                sheet1.cell(row=row_idx, column=3).value = height
                sheet1.cell(row=row_idx, column=4).value = obj['class']
                sheet1.cell(row=row_idx, column=5).value = obj['x']
                sheet1.cell(row=row_idx, column=6).value = obj['y']

    print(xlsx_file_name)
    write_wb.save(filename=xlsx_file_name)
    # / 2. 추출한 값들로 엑셀데이터 만들기

    # 3. csv 변환 후 기존 엑셀데이터 삭제
    data_xls = pd.read_excel(xlsx_file_name, 'Sheet', index_col=None)
    data_xls.to_csv(EXT_PATH + FOLDER_NAME + '.csv', encoding='utf-8')
    os.remove(xlsx_file_name)
    # / 3. csv 변환 후 기존 엑셀데이터 삭제

##### 4차 끝 #####

# select U.store_file_path
#
# from  TB_PRJ_DATA     D left join TB_MEMBER M on M.member_id = D.work_user
#
# left join TB_TMPL_IMG_LABEL_DATA L on D.data_idx  = L.data_idx
# left join CW_SOURCE S on S.source_id = L.src_idx
# left join CW_FILE_UNPACKED U  on U.file_id = S.file_id
#
#
# where D.prj_idx in () and D.prog_state_cd="ALL_FINISHED";
